from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Any
import uuid
import json
import re
import io
import csv
from datetime import datetime, timezone
from openai import AsyncOpenAI

try:
    from emergentintegrations.llm.openai import OpenAIChatRealtime
    HAS_REALTIME = True
except ImportError:
    OpenAIChatRealtime = None
    HAS_REALTIME = False

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# OpenAI client - use OPENAI_API_KEY if provided, otherwise fall back to EMERGENT_LLM_KEY
def get_openai_client():
    api_key = os.environ.get('OPENAI_API_KEY') or os.environ.get('EMERGENT_LLM_KEY')
    if not api_key:
        raise ValueError("No API key configured. Set OPENAI_API_KEY in .env")
    return AsyncOpenAI(api_key=api_key)

# Model for chat and vision (gpt-4o requires higher tier; gpt-4o-mini is widely available)
OPENAI_CHAT_MODEL = os.environ.get("OPENAI_CHAT_MODEL", "gpt-4o-mini")


def _extract_chart_json(text: str) -> Optional[dict]:
    """Extract a chart JSON object from model response. Expects format: {"chart_type": "bar", "title": "...", "data": [{"category": "X", "count": N}, ...]}"""
    if "chart_type" not in text.lower() or '"data"' not in text:
        return None
    # Find start of JSON object containing chart_type (allow nested braces)
    start = text.find('{"chart_type"')
    if start == -1:
        start = text.find("{\"chart_type\"")
    if start == -1:
        return None
    depth = 0
    for i in range(start, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                try:
                    obj = json.loads(text[start : i + 1])
                    if isinstance(obj.get("data"), list) and obj.get("chart_type"):
                        # Normalize to bar chart format: category + count (frontend expects these)
                        normalized = []
                        for item in obj["data"]:
                            if not isinstance(item, dict):
                                continue
                            cat = item.get("category") or item.get("name") or str(len(normalized))
                            val = item.get("count") if "count" in item else item.get("value", 0)
                            try:
                                val = int(float(val))
                            except (TypeError, ValueError):
                                val = 0
                            normalized.append({"category": str(cat)[:30], "count": val})
                        if normalized:
                            obj["data"] = normalized[:12]
                            obj["chart_type"] = "bar"
                            obj["title"] = obj.get("title") or "Chart"
                            return obj
                except (json.JSONDecodeError, TypeError):
                    pass
                return None
    return None

# Initialize OpenAI Realtime Chat for WebRTC (optional; requires emergentintegrations)
openai_api_key = os.environ.get('OPENAI_API_KEY') or os.environ.get('EMERGENT_LLM_KEY')
realtime_chat = OpenAIChatRealtime(api_key=openai_api_key) if (HAS_REALTIME and openai_api_key) else None

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Create a router for realtime API and register OpenAI Realtime routes
realtime_router = APIRouter()
if realtime_chat:
    OpenAIChatRealtime.register_openai_realtime_router(realtime_router, realtime_chat)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# --------------------- Models ---------------------

class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

class Finding(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    timestamp: str
    severity: str  # HIGH, MEDIUM, LOW
    title: str
    recommendation: str
    confidence: float
    category: str

class ChecklistItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category: str
    item: str
    result: str  # PASS, FAIL, MONITOR
    severity: str
    evidence: Optional[str] = None
    recommended_action: Optional[str] = None
    confidence: float

class PartMatch(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    part_number: str
    part_name: str
    fitment_certainty: float
    compatible_models: List[str]

class MediaItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: str  # photo, video
    url: str
    thumbnail: Optional[str] = None
    timestamp: str
    caption: Optional[str] = None

class Inspection(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    equipment_model: str
    serial_number: str
    customer: str
    location: str
    inspection_type: str  # Daily Walkaround, Safety, TA1
    status: str  # Draft, In Progress, Submitted, PASS, FAIL, MONITOR
    date: str
    inspector: str
    summary: Optional[str] = None
    safety_findings: Optional[List[str]] = []
    action_items: Optional[List[dict]] = []
    findings: Optional[List[Finding]] = []
    checklist: Optional[List[ChecklistItem]] = []
    parts_matches: Optional[List[PartMatch]] = []
    media: Optional[List[MediaItem]] = []
    similar_inspections: Optional[List[dict]] = []

class InspectionCreate(BaseModel):
    equipment_model: str
    serial_number: str
    customer: str
    location: str
    inspection_type: str

class ChatMessage(BaseModel):
    role: str  # user, assistant
    content: str
    timestamp: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    chart_data: Optional[dict] = None

class ChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = None

class ChatResponse(BaseModel):
    response: str
    chart_data: Optional[dict] = None

# Session-scoped uploaded documents for chatbot context (session_id -> {filename, text})
SESSION_DOCUMENTS: dict[str, dict[str, Any]] = {}
MAX_DOCUMENT_CHARS = 50_000  # truncate to stay within context window

def extract_text_from_pdf(content: bytes) -> str:
    """Extract text from PDF bytes. Returns empty string on failure."""
    try:
        from pypdf import PdfReader
        from io import BytesIO
        reader = PdfReader(BytesIO(content))
        parts = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                parts.append(t)
        text = "\n\n".join(parts).strip()
        if len(text) > MAX_DOCUMENT_CHARS:
            text = text[:MAX_DOCUMENT_CHARS] + "\n\n[Document truncated for length.]"
        return text
    except Exception as e:
        logger.warning(f"PDF extraction failed: {e}")
        return ""

class AnalyticsData(BaseModel):
    failed_parts: List[dict]
    inspections_over_time: List[dict]
    pass_fail_monitor: dict

# AI Vision Analysis Request
class VisionAnalysisRequest(BaseModel):
    image_base64: str
    context: Optional[str] = "equipment inspection"

class VisionAnalysisResponse(BaseModel):
    analysis: str
    findings: List[dict]
    severity: str
    should_alert: bool

# Text to Speech Request
class TTSRequest(BaseModel):
    text: str
    voice: Optional[str] = "alloy"

# Speech to Text Request  
class STTRequest(BaseModel):
    audio_base64: str

# --------------------- Mock Data ---------------------

MOCK_INSPECTIONS = [
    {
        "id": "insp-001",
        "equipment_model": "CAT 320 Excavator",
        "serial_number": "CAT0320X12345",
        "customer": "BuildCo Industries",
        "location": "Dallas, TX",
        "inspection_type": "Daily Walkaround",
        "status": "PASS",
        "date": "2025-01-15",
        "inspector": "Sriram N.",
        "summary": "Equipment in excellent condition. All systems operational.",
        "safety_findings": [],
        "action_items": []
    },
    {
        "id": "insp-002",
        "equipment_model": "CAT D6 Dozer",
        "serial_number": "CAT0D6X67890",
        "customer": "Highway Construction LLC",
        "location": "Austin, TX",
        "inspection_type": "Safety",
        "status": "FAIL",
        "date": "2025-01-14",
        "inspector": "Sriram N.",
        "summary": "Critical hydraulic leak detected in main boom cylinder. Immediate repair required.",
        "safety_findings": ["Hydraulic leak - High pressure line compromised"],
        "action_items": [
            {"priority": 1, "action": "Replace hydraulic line", "risk": "High - Equipment failure risk"},
            {"priority": 2, "action": "Check all fluid levels", "risk": "Medium"}
        ]
    },
    {
        "id": "insp-003",
        "equipment_model": "CAT 966 Wheel Loader",
        "serial_number": "CAT0966X11111",
        "customer": "Quarry Masters Inc",
        "location": "Houston, TX",
        "inspection_type": "TA1",
        "status": "MONITOR",
        "date": "2025-01-13",
        "inspector": "Sriram N.",
        "summary": "Minor wear on bucket teeth. Schedule replacement within 30 days.",
        "safety_findings": [],
        "action_items": [
            {"priority": 1, "action": "Order replacement bucket teeth", "risk": "Low - Wear item"},
            {"priority": 2, "action": "Re-inspect in 2 weeks", "risk": "Low"}
        ]
    },
    {
        "id": "insp-004",
        "equipment_model": "CAT 745 Articulated Truck",
        "serial_number": "CAT0745X22222",
        "customer": "Mountain Mining Co",
        "location": "Denver, CO",
        "inspection_type": "Daily Walkaround",
        "status": "PASS",
        "date": "2025-01-12",
        "inspector": "Sriram N.",
        "summary": "All systems nominal. Tire pressure within spec.",
        "safety_findings": [],
        "action_items": []
    },
    {
        "id": "insp-005",
        "equipment_model": "CAT 336 Excavator",
        "serial_number": "CAT0336X33333",
        "customer": "Urban Development Corp",
        "location": "Phoenix, AZ",
        "inspection_type": "Safety",
        "status": "In Progress",
        "date": "2025-01-11",
        "inspector": "Sriram N.",
        "summary": "",
        "safety_findings": [],
        "action_items": []
    }
]

MOCK_ANALYTICS = {
    "failed_parts": [
        {"category": "Hydraulics", "count": 12, "percentage": 35},
        {"category": "Engine", "count": 8, "percentage": 23},
        {"category": "Electrical", "count": 6, "percentage": 17},
        {"category": "Undercarriage", "count": 5, "percentage": 15},
        {"category": "Attachments", "count": 3, "percentage": 10}
    ],
    "inspections_over_time": [
        {"month": "Aug", "count": 42},
        {"month": "Sep", "count": 38},
        {"month": "Oct", "count": 55},
        {"month": "Nov", "count": 47},
        {"month": "Dec", "count": 52},
        {"month": "Jan", "count": 31}
    ],
    "pass_fail_monitor": {
        "pass": 156,
        "fail": 23,
        "monitor": 45
    }
}

# Detailed analytics per category (for category drill-down)
def _category_analytics(category: str) -> dict:
    """Mock detailed analytics for a category. Varies by category name."""
    base = {
        "Hydraulics": {
            "failures_over_time": [
                {"month": "Aug", "count": 3},
                {"month": "Sep", "count": 2},
                {"month": "Oct", "count": 4},
                {"month": "Nov", "count": 1},
                {"month": "Dec", "count": 2},
                {"month": "Jan", "count": 0},
            ],
            "item_breakdown": [
                {"item": "Main Boom Cylinder", "pass": 8, "fail": 5, "monitor": 2},
                {"item": "Stick Cylinder", "pass": 12, "fail": 1, "monitor": 2},
                {"item": "Hydraulic Hoses", "pass": 6, "fail": 4, "monitor": 5},
                {"item": "Pump & Reservoir", "pass": 10, "fail": 2, "monitor": 3},
            ],
            "severity_breakdown": [{"severity": "HIGH", "count": 5}, {"severity": "MEDIUM", "count": 4}, {"severity": "LOW", "count": 3}],
            "top_recommended_actions": [
                {"action": "Replace high-pressure line", "count": 6},
                {"action": "Inspect seals and fittings", "count": 4},
                {"action": "Top off fluid / check for leaks", "count": 2},
            ],
            "recent_inspections": [
                {"id": "insp-002", "equipment": "CAT D6 Dozer", "date": "2025-01-14", "result": "FAIL", "summary": "Hydraulic leak on main boom cylinder"},
                {"id": "insp-005", "equipment": "CAT 320 Excavator", "date": "2025-01-10", "result": "FAIL", "summary": "Stick cylinder seal wear"},
                {"id": "insp-008", "equipment": "CAT D6 Dozer", "date": "2025-01-05", "result": "MONITOR", "summary": "Minor hose wear observed"},
            ],
            "heatmap_global": [
                {"id": "USA", "topo_id": "840", "name": "United States", "high": 4, "medium": 3, "low": 2, "severity_index": 0.72},
                {"id": "CAN", "topo_id": "124", "name": "Canada", "high": 1, "medium": 2, "low": 1, "severity_index": 0.55},
                {"id": "GBR", "topo_id": "826", "name": "United Kingdom", "high": 2, "medium": 2, "low": 1, "severity_index": 0.68},
                {"id": "DEU", "topo_id": "276", "name": "Germany", "high": 1, "medium": 1, "low": 2, "severity_index": 0.42},
                {"id": "BRA", "topo_id": "076", "name": "Brazil", "high": 2, "medium": 1, "low": 0, "severity_index": 0.78},
                {"id": "AUS", "topo_id": "036", "name": "Australia", "high": 1, "medium": 2, "low": 1, "severity_index": 0.52},
                {"id": "IND", "topo_id": "356", "name": "India", "high": 2, "medium": 2, "low": 1, "severity_index": 0.65},
                {"id": "MEX", "topo_id": "484", "name": "Mexico", "high": 1, "medium": 1, "low": 1, "severity_index": 0.48},
            ],
            "heatmap_local": {
                "USA": [
                    {"id": "TX", "name": "Texas", "high": 2, "medium": 1, "low": 0, "severity_index": 0.85},
                    {"id": "CA", "name": "California", "high": 1, "medium": 1, "low": 1, "severity_index": 0.58},
                    {"id": "IL", "name": "Illinois", "high": 1, "medium": 1, "low": 0, "severity_index": 0.72},
                    {"id": "AZ", "name": "Arizona", "high": 0, "medium": 1, "low": 1, "severity_index": 0.45},
                    {"id": "CO", "name": "Colorado", "high": 0, "medium": 0, "low": 1, "severity_index": 0.28},
                ],
                "CAN": [{"id": "ON", "name": "Ontario", "high": 0, "medium": 1, "low": 1, "severity_index": 0.4}, {"id": "AB", "name": "Alberta", "high": 1, "medium": 0, "low": 0, "severity_index": 0.9}],
                "GBR": [{"id": "ENG", "name": "England", "high": 1, "medium": 1, "low": 0, "severity_index": 0.7}, {"id": "SCT", "name": "Scotland", "high": 0, "medium": 1, "low": 0, "severity_index": 0.5}],
            },
            "insight_summary": "Hydraulics account for the highest share of failures. Main Boom Cylinder and hose assemblies are the most common failure points. Consider scheduling preventive hose replacement.",
        },
        "Engine": {
            "failures_over_time": [
                {"month": "Aug", "count": 1},
                {"month": "Sep", "count": 2},
                {"month": "Oct", "count": 2},
                {"month": "Nov", "count": 1},
                {"month": "Dec", "count": 1},
                {"month": "Jan", "count": 1},
            ],
            "item_breakdown": [
                {"item": "Oil Level / Leaks", "pass": 10, "fail": 3, "monitor": 2},
                {"item": "Coolant Level", "pass": 12, "fail": 1, "monitor": 2},
                {"item": "Air Filter", "pass": 8, "fail": 2, "monitor": 5},
                {"item": "Belts & Hoses", "pass": 11, "fail": 2, "monitor": 2},
            ],
            "severity_breakdown": [{"severity": "HIGH", "count": 2}, {"severity": "MEDIUM", "count": 3}, {"severity": "LOW", "count": 3}],
            "top_recommended_actions": [
                {"action": "Replace oil filter / fix leak", "count": 4},
                {"action": "Top off coolant", "count": 2},
                {"action": "Replace air filter", "count": 2},
            ],
            "recent_inspections": [
                {"id": "insp-003", "equipment": "CAT 320 Excavator", "date": "2025-01-12", "result": "FAIL", "summary": "Engine oil leak at filter housing"},
                {"id": "insp-006", "equipment": "CAT D6 Dozer", "date": "2025-01-08", "result": "MONITOR", "summary": "Coolant level at minimum"},
            ],
            "heatmap_global": [
                {"id": "USA", "topo_id": "840", "name": "United States", "high": 2, "medium": 2, "low": 3, "severity_index": 0.48},
                {"id": "CAN", "topo_id": "124", "name": "Canada", "high": 1, "medium": 1, "low": 1, "severity_index": 0.5},
                {"id": "MEX", "topo_id": "484", "name": "Mexico", "high": 1, "medium": 1, "low": 0, "severity_index": 0.65},
            ],
            "heatmap_local": {"USA": [{"id": "TX", "name": "Texas", "high": 1, "medium": 0, "low": 1, "severity_index": 0.55}, {"id": "CA", "name": "California", "high": 0, "medium": 1, "low": 1, "severity_index": 0.35}]},
            "insight_summary": "Engine-related issues are mostly oil and coolant. Oil leaks and filter condition are the top drivers. Regular fluid checks can reduce failures.",
        },
        "Electrical": {
            "failures_over_time": [
                {"month": "Aug", "count": 2},
                {"month": "Sep", "count": 1},
                {"month": "Oct", "count": 1},
                {"month": "Nov", "count": 1},
                {"month": "Dec", "count": 0},
                {"month": "Jan", "count": 1},
            ],
            "item_breakdown": [
                {"item": "Battery & Connections", "pass": 9, "fail": 2, "monitor": 4},
                {"item": "Wiring & Harness", "pass": 10, "fail": 2, "monitor": 3},
                {"item": "Sensors", "pass": 11, "fail": 1, "monitor": 3},
                {"item": "Lighting", "pass": 13, "fail": 1, "monitor": 1},
            ],
            "severity_breakdown": [{"severity": "HIGH", "count": 1}, {"severity": "MEDIUM", "count": 2}, {"severity": "LOW", "count": 3}],
            "top_recommended_actions": [
                {"action": "Clean battery terminals / load test", "count": 3},
                {"action": "Repair or replace damaged wiring", "count": 2},
            ],
            "recent_inspections": [
                {"id": "insp-004", "equipment": "CAT D6 Dozer", "date": "2025-01-11", "result": "FAIL", "summary": "Battery voltage low; alternator output marginal"},
            ],
            "heatmap_global": [
                {"id": "USA", "topo_id": "840", "name": "United States", "high": 1, "medium": 1, "low": 2, "severity_index": 0.45},
                {"id": "DEU", "topo_id": "276", "name": "Germany", "high": 0, "medium": 1, "low": 1, "severity_index": 0.38},
            ],
            "heatmap_local": {"USA": [{"id": "TX", "name": "Texas", "high": 1, "medium": 0, "low": 0, "severity_index": 0.85}]},
            "insight_summary": "Electrical failures are less frequent but often battery or wiring related. Load testing batteries during inspections can catch issues early.",
        },
        "Undercarriage": {
            "failures_over_time": [
                {"month": "Aug", "count": 1},
                {"month": "Sep", "count": 1},
                {"month": "Oct", "count": 1},
                {"month": "Nov", "count": 1},
                {"month": "Dec", "count": 0},
                {"month": "Jan", "count": 1},
            ],
            "item_breakdown": [
                {"item": "Track Tension", "pass": 8, "fail": 1, "monitor": 6},
                {"item": "Rollers & Idlers", "pass": 10, "fail": 2, "monitor": 3},
                {"item": "Sprockets", "pass": 11, "fail": 1, "monitor": 3},
                {"item": "Track Pads", "pass": 9, "fail": 1, "monitor": 5},
            ],
            "severity_breakdown": [{"severity": "HIGH", "count": 1}, {"severity": "MEDIUM", "count": 2}, {"severity": "LOW", "count": 2}],
            "top_recommended_actions": [
                {"action": "Adjust track tension", "count": 3},
                {"action": "Replace worn rollers", "count": 2},
            ],
            "recent_inspections": [
                {"id": "insp-007", "equipment": "CAT D6 Dozer", "date": "2025-01-07", "result": "MONITOR", "summary": "Track tension at upper limit; re-check in 2 weeks"},
            ],
            "heatmap_global": [{"id": "USA", "topo_id": "840", "name": "United States", "high": 0, "medium": 1, "low": 2, "severity_index": 0.35}, {"id": "CAN", "topo_id": "124", "name": "Canada", "high": 0, "medium": 1, "low": 0, "severity_index": 0.5}],
            "heatmap_local": {"USA": [{"id": "TX", "name": "Texas", "high": 0, "medium": 1, "low": 1, "severity_index": 0.4}]},
            "insight_summary": "Undercarriage issues are often tension and wear. Regular tension checks and pad wear monitoring help avoid unexpected downtime.",
        },
        "Attachments": {
            "failures_over_time": [
                {"month": "Aug", "count": 0},
                {"month": "Sep", "count": 1},
                {"month": "Oct", "count": 1},
                {"month": "Nov", "count": 0},
                {"month": "Dec", "count": 1},
                {"month": "Jan", "count": 0},
            ],
            "item_breakdown": [
                {"item": "Bucket / Blade Condition", "pass": 12, "fail": 1, "monitor": 2},
                {"item": "Pins & Bushings", "pass": 10, "fail": 2, "monitor": 3},
                {"item": "Cutting Edges", "pass": 11, "fail": 1, "monitor": 3},
            ],
            "severity_breakdown": [{"severity": "HIGH", "count": 0}, {"severity": "MEDIUM", "count": 1}, {"severity": "LOW", "count": 2}],
            "top_recommended_actions": [
                {"action": "Replace worn pins and bushings", "count": 2},
                {"action": "Replace cutting edges", "count": 1},
            ],
            "recent_inspections": [
                {"id": "insp-009", "equipment": "CAT 320 Excavator", "date": "2025-01-04", "result": "MONITOR", "summary": "Bucket teeth wear; schedule replacement"},
            ],
            "heatmap_global": [{"id": "USA", "topo_id": "840", "name": "United States", "high": 0, "medium": 1, "low": 2, "severity_index": 0.32}],
            "heatmap_local": {"USA": [{"id": "AZ", "name": "Arizona", "high": 0, "medium": 1, "low": 0, "severity_index": 0.5}]},
            "insight_summary": "Attachment failures are the lowest. Focus on pins, bushings, and cutting edges to extend attachment life.",
        },
    }
    data = base.get(category)
    if not data:
        # Default template for unknown category
        data = {
            "failures_over_time": MOCK_ANALYTICS["inspections_over_time"],
            "item_breakdown": [{"item": "Item A", "pass": 5, "fail": 2, "monitor": 1}, {"item": "Item B", "pass": 6, "fail": 1, "monitor": 0}],
            "severity_breakdown": [{"severity": "HIGH", "count": 0}, {"severity": "MEDIUM", "count": 1}, {"severity": "LOW", "count": 1}],
            "top_recommended_actions": [{"action": "Review and repair", "count": 1}],
            "recent_inspections": [],
            "heatmap_global": [{"id": "USA", "topo_id": "840", "name": "United States", "high": 0, "medium": 1, "low": 0, "severity_index": 0.5}],
            "heatmap_local": {},
            "insight_summary": f"Inspection data for {category}.",
        }
    return data

MOCK_INSPECTION_DETAIL = {
    "id": "insp-002",
    "equipment_model": "CAT D6 Dozer",
    "serial_number": "CAT0D6X67890",
    "customer": "Highway Construction LLC",
    "location": "Austin, TX",
    "inspection_type": "Safety",
    "status": "FAIL",
    "date": "2025-01-14",
    "inspector": "Sriram N.",
    "summary": "This safety inspection identified a critical hydraulic leak in the main boom cylinder that requires immediate attention. The leak was detected during visual inspection and confirmed with pressure testing. All other systems are operating within normal parameters, but the equipment should be taken out of service until repairs are completed.",
    "safety_findings": [
        "Critical: Hydraulic leak detected in main boom cylinder - High pressure line compromised",
        "Warning: Operator visibility reduced due to cracked side mirror"
    ],
    "action_items": [
        {"priority": 1, "action": "Replace hydraulic high-pressure line on main boom cylinder", "risk": "Critical - Equipment failure and safety hazard", "why": "Leak can cause sudden loss of boom control"},
        {"priority": 2, "action": "Replace cracked side mirror", "risk": "Medium - Reduced operator visibility", "why": "Safety compliance requirement"},
        {"priority": 3, "action": "Schedule follow-up inspection after repairs", "risk": "Low", "why": "Verify repairs and clear equipment for operation"}
    ],
    "findings": [
        {"id": "f1", "timestamp": "10:23:45", "severity": "HIGH", "title": "Hydraulic Leak Detected", "recommendation": "Immediately shut down and replace high-pressure line", "confidence": 0.95, "category": "Hydraulics"},
        {"id": "f2", "timestamp": "10:25:12", "severity": "MEDIUM", "title": "Side Mirror Cracked", "recommendation": "Replace mirror before next shift", "confidence": 0.98, "category": "Safety Equipment"},
        {"id": "f3", "timestamp": "10:28:33", "severity": "LOW", "title": "Minor Rust on Step Rails", "recommendation": "Schedule touch-up paint during next service", "confidence": 0.87, "category": "Structural"}
    ],
    "checklist": [
        {"id": "c1", "category": "Hydraulics", "item": "Main Boom Cylinder", "result": "FAIL", "severity": "HIGH", "evidence": "photo_001.jpg", "recommended_action": "Replace high-pressure line", "confidence": 0.95},
        {"id": "c2", "category": "Hydraulics", "item": "Stick Cylinder", "result": "PASS", "severity": "LOW", "evidence": None, "recommended_action": None, "confidence": 0.92},
        {"id": "c3", "category": "Engine", "item": "Oil Level", "result": "PASS", "severity": "LOW", "evidence": None, "recommended_action": None, "confidence": 0.99},
        {"id": "c4", "category": "Engine", "item": "Coolant Level", "result": "PASS", "severity": "LOW", "evidence": None, "recommended_action": None, "confidence": 0.98},
        {"id": "c5", "category": "Safety Equipment", "item": "Side Mirrors", "result": "FAIL", "severity": "MEDIUM", "evidence": "photo_002.jpg", "recommended_action": "Replace cracked mirror", "confidence": 0.98},
        {"id": "c6", "category": "Safety Equipment", "item": "Backup Camera", "result": "PASS", "severity": "LOW", "evidence": None, "recommended_action": None, "confidence": 0.96},
        {"id": "c7", "category": "Undercarriage", "item": "Track Tension", "result": "MONITOR", "severity": "LOW", "evidence": None, "recommended_action": "Re-check in 2 weeks", "confidence": 0.85},
        {"id": "c8", "category": "Structural", "item": "Step Rails", "result": "MONITOR", "severity": "LOW", "evidence": "photo_003.jpg", "recommended_action": "Touch-up paint", "confidence": 0.87}
    ],
    "parts_matches": [
        {"id": "p1", "part_number": "5I-4461", "part_name": "Hydraulic Hose Assembly", "fitment_certainty": 0.97, "compatible_models": ["D6", "D6T", "D6N"]},
        {"id": "p2", "part_number": "1U-1857", "part_name": "O-Ring Seal Kit", "fitment_certainty": 0.94, "compatible_models": ["D6", "D6T", "D7"]},
        {"id": "p3", "part_number": "9W-3214", "part_name": "Side Mirror Assembly", "fitment_certainty": 0.99, "compatible_models": ["D6", "D6T", "D6N", "D6R"]}
    ],
    "media": [
        {"id": "m1", "type": "photo", "url": "https://images.unsplash.com/photo-1581094288338-2314dddb7ece?w=800", "thumbnail": "https://images.unsplash.com/photo-1581094288338-2314dddb7ece?w=200", "timestamp": "10:23:45", "caption": "Hydraulic leak on main boom cylinder"},
        {"id": "m2", "type": "photo", "url": "https://images.unsplash.com/photo-1504307651254-35680f356dfd?w=800", "thumbnail": "https://images.unsplash.com/photo-1504307651254-35680f356dfd?w=200", "timestamp": "10:25:12", "caption": "Cracked side mirror"},
        {"id": "m3", "type": "photo", "url": "https://images.unsplash.com/photo-1566041510639-8d95a2490bfb?w=800", "thumbnail": "https://images.unsplash.com/photo-1566041510639-8d95a2490bfb?w=200", "timestamp": "10:28:33", "caption": "Minor rust on step rails"},
        {"id": "m4", "type": "video", "url": "https://images.unsplash.com/photo-1621905252507-b35492cc74b4?w=800", "thumbnail": "https://images.unsplash.com/photo-1621905252507-b35492cc74b4?w=200", "timestamp": "10:30:00", "caption": "Equipment walkthrough video"}
    ],
    "similar_inspections": [
        {"id": "insp-010", "title": "Similar Hydraulic Issues Cluster", "summary": "3 other D6 units in the fleet have shown similar hydraulic line wear in the past 90 days", "count": 3},
        {"id": "insp-011", "title": "Mirror Damage Pattern", "summary": "5 units reported side mirror damage this quarter, possible site condition issue", "count": 5}
    ]
}

# --------------------- Routes ---------------------

@api_router.get("/")
async def root():
    return {"message": "Cat Inspect AI API"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    _ = await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    return status_checks

# Inspections endpoints
@api_router.get("/inspections")
async def get_inspections(status: Optional[str] = None, inspection_type: Optional[str] = None, search: Optional[str] = None):
    """Get list of inspections with optional filters"""
    results = MOCK_INSPECTIONS.copy()
    
    if status and status != "all":
        results = [i for i in results if i["status"].lower() == status.lower()]
    
    if inspection_type and inspection_type != "all":
        results = [i for i in results if i["inspection_type"].lower() == inspection_type.lower()]
    
    if search:
        search_lower = search.lower()
        results = [i for i in results if 
                   search_lower in i["equipment_model"].lower() or 
                   search_lower in i["serial_number"].lower() or
                   search_lower in i["customer"].lower() or
                   search_lower in i["location"].lower()]
    
    return results

# Store for dynamically created inspections
CREATED_INSPECTIONS = {}

@api_router.get("/inspections/{inspection_id}")
async def get_inspection(inspection_id: str):
    """Get single inspection detail"""
    # Return detailed mock for insp-002, otherwise return basic mock
    if inspection_id == "insp-002":
        return MOCK_INSPECTION_DETAIL
    
    # Check if it's a dynamically created inspection
    if inspection_id in CREATED_INSPECTIONS:
        return CREATED_INSPECTIONS[inspection_id]
    
    for insp in MOCK_INSPECTIONS:
        if insp["id"] == inspection_id:
            return {**insp, **{
                "findings": MOCK_INSPECTION_DETAIL["findings"],
                "checklist": MOCK_INSPECTION_DETAIL["checklist"],
                "parts_matches": MOCK_INSPECTION_DETAIL["parts_matches"],
                "media": MOCK_INSPECTION_DETAIL["media"],
                "similar_inspections": MOCK_INSPECTION_DETAIL["similar_inspections"]
            }}
    
    # For any unknown ID, return a mock completed inspection
    return {
        "id": inspection_id,
        "equipment_model": "CAT D6 Dozer",
        "serial_number": "CAT0D6X" + inspection_id[-5:],
        "customer": "New Customer",
        "location": "Dallas, TX",
        "inspection_type": "Safety",
        "status": "Submitted",
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "inspector": "Sriram N.",
        "summary": "This inspection has been completed successfully. The AI assistant analyzed the equipment and identified several items for review. All critical safety checks passed, with minor maintenance recommendations noted below.",
        "safety_findings": ["No critical safety issues detected"],
        "action_items": [
            {"priority": 1, "action": "Review captured findings", "risk": "Low", "why": "Ensure all items documented"},
            {"priority": 2, "action": "Schedule follow-up if needed", "risk": "Low", "why": "Preventive maintenance"}
        ],
        "findings": MOCK_INSPECTION_DETAIL["findings"],
        "checklist": MOCK_INSPECTION_DETAIL["checklist"],
        "parts_matches": MOCK_INSPECTION_DETAIL["parts_matches"],
        "media": MOCK_INSPECTION_DETAIL["media"],
        "similar_inspections": MOCK_INSPECTION_DETAIL["similar_inspections"]
    }

@api_router.post("/inspections")
async def create_inspection(inspection: InspectionCreate):
    """Create a new inspection"""
    new_id = f"insp-{str(uuid.uuid4())[:8]}"
    new_inspection = {
        "id": new_id,
        "equipment_model": inspection.equipment_model,
        "serial_number": inspection.serial_number,
        "customer": inspection.customer,
        "location": inspection.location,
        "inspection_type": inspection.inspection_type,
        "status": "In Progress",
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "inspector": "Sriram N.",
        "summary": "",
        "safety_findings": [],
        "action_items": []
    }
    return new_inspection

@api_router.put("/inspections/{inspection_id}/checklist/{item_id}")
async def update_checklist_item(inspection_id: str, item_id: str, result: str):
    """Update a checklist item result (inspector override)"""
    return {"success": True, "item_id": item_id, "new_result": result}

@api_router.post("/inspections/{inspection_id}/finish")
async def finish_inspection(inspection_id: str):
    """Finish inspection and generate report"""
    return {
        "success": True,
        "inspection_id": inspection_id,
        "status": "Submitted",
        "report_generated": True
    }

# Analytics endpoint
@api_router.get("/analytics")
async def get_analytics():
    """Get analytics data for dashboard"""
    return MOCK_ANALYTICS


@api_router.get("/analytics/category/{category}")
async def get_analytics_by_category(category: str):
    """Get detailed analytics for a specific category (e.g. Hydraulics, Engine)."""
    name_in_url = (category or "").strip()
    category_key = next(
        (c["category"] for c in MOCK_ANALYTICS["failed_parts"] if c["category"].lower() == name_in_url.lower()),
        name_in_url or "Other",
    )
    detail = _category_analytics(category_key)
    part_row = next((p for p in MOCK_ANALYTICS["failed_parts"] if p["category"].lower() == category_key.lower()), None)
    return {
        "category": category_key,
        "total_failures": part_row["count"] if part_row else 0,
        "percentage_of_all": part_row["percentage"] if part_row else 0,
        **detail,
    }


# ---------- Export: PDF (single report) and CSV (all + analytics) ----------
def _build_inspection_pdf(detail: dict) -> bytes:
    """Build a PDF report for one inspection using reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name="Title", parent=styles["Heading1"], fontSize=16, spaceAfter=12)
    h2_style = ParagraphStyle(name="H2", parent=styles["Heading2"], fontSize=12, spaceAfter=6, spaceBefore=12)

    story = []
    story.append(Paragraph("CAT Inspect – Inspection Report", title_style))
    story.append(Spacer(1, 0.2 * inch))

    # Meta (ensure no None for reportlab)
    def _s(v):
        return str(v).strip() if v is not None else ""

    meta = [
        ["Equipment", _s(detail.get("equipment_model"))],
        ["Serial", _s(detail.get("serial_number"))],
        ["Customer", _s(detail.get("customer"))],
        ["Location", _s(detail.get("location"))],
        ["Date", _s(detail.get("date"))],
        ["Type", _s(detail.get("inspection_type"))],
        ["Status", _s(detail.get("status"))],
        ["Inspector", _s(detail.get("inspector"))],
    ]
    t = Table(meta, colWidths=[1.2 * inch, 4.3 * inch])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (0, -1), 9),
        ("FONTSIZE", (1, 0), (1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.25 * inch))

    summary = (detail.get("summary") or "").strip()
    if summary:
        story.append(Paragraph("Summary", h2_style))
        story.append(Paragraph(summary.replace("\n", "<br/>"), styles["Normal"]))
        story.append(Spacer(1, 0.15 * inch))

    safety = detail.get("safety_findings") or []
    if safety:
        story.append(Paragraph("Safety findings", h2_style))
        for s in safety:
            story.append(Paragraph(f"• {_s(s)}", styles["Normal"]))
        story.append(Spacer(1, 0.15 * inch))

    actions = detail.get("action_items") or []
    if actions:
        story.append(Paragraph("Action items", h2_style))
        rows = [["Priority", "Action", "Risk", "Reason"]]
        for a in actions:
            rows.append([
                str(a.get("priority") if a.get("priority") is not None else ""),
                _s(a.get("action")),
                _s(a.get("risk")),
                _s(a.get("why")),
            ])
        t2 = Table(rows, colWidths=[0.6 * inch, 2.2 * inch, 1.4 * inch, 1.3 * inch])
        t2.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F7B500")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(t2)
        story.append(Spacer(1, 0.2 * inch))

    checklist = detail.get("checklist") or []
    if checklist:
        story.append(Paragraph("Checklist", h2_style))
        rows = [["Category", "Item", "Result", "Severity", "Recommended action"]]
        for c in checklist:
            rows.append([
                _s(c.get("category")),
                _s(c.get("item")),
                _s(c.get("result")),
                _s(c.get("severity")),
                _s(c.get("recommended_action")),
            ])
        t3 = Table(rows, colWidths=[1.1 * inch, 1.8 * inch, 0.7 * inch, 0.7 * inch, 1.2 * inch])
        t3.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t3)

    doc.build(story)
    return buf.getvalue()


@api_router.get("/export/inspection/{inspection_id}/pdf")
async def export_inspection_pdf(inspection_id: str):
    """Export a single inspection report as PDF."""
    detail = await get_inspection(inspection_id)
    pdf_bytes = _build_inspection_pdf(detail)
    filename = f"inspection-report-{inspection_id}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@api_router.get("/export/all")
async def export_all_csv():
    """Export all inspections and analytics as a single CSV for Google Sheets. Includes sections for charts."""
    inspections = await get_inspections()
    analytics = MOCK_ANALYTICS
    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["CAT Inspect – Full Export (CSV)"])
    w.writerow([datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    w.writerow(["Import into Google Sheets, then use Insert > Chart on any data block below to create graphs."])
    w.writerow([])

    w.writerow(["INSPECTIONS (detail)"])
    w.writerow([
        "ID", "Equipment", "Serial", "Customer", "Location", "Date", "Type", "Status", "Inspector",
        "Safety findings count", "Action items count", "Summary"
    ])
    for i in inspections:
        safety = i.get("safety_findings") or []
        actions = i.get("action_items") or []
        w.writerow([
            i.get("id", ""),
            i.get("equipment_model", ""),
            i.get("serial_number", ""),
            i.get("customer", ""),
            i.get("location", ""),
            i.get("date", ""),
            i.get("inspection_type", ""),
            i.get("status", ""),
            i.get("inspector", ""),
            len(safety),
            len(actions),
            ((i.get("summary") or "")[:300]),
        ])
    w.writerow([])

    w.writerow(["ANALYTICS – Failed parts by category (Bar chart: Category = X, Count = Y)"])
    w.writerow(["Category", "Count", "Percentage"])
    for p in (analytics.get("failed_parts") or []):
        w.writerow([p.get("category", ""), p.get("count", ""), p.get("percentage", "")])
    w.writerow([])

    w.writerow(["ANALYTICS – Inspections over time (Line chart: Month = X, Count = Y)"])
    w.writerow(["Month", "Count"])
    for r in (analytics.get("inspections_over_time") or []):
        w.writerow([r.get("month", ""), r.get("count", "")])
    w.writerow([])

    w.writerow(["ANALYTICS – Pass / Fail / Monitor (Pie chart: Outcome = labels, Count = values)"])
    w.writerow(["Outcome", "Count"])
    pfm = analytics.get("pass_fail_monitor") or {}
    for label, key in [("Pass", "pass"), ("Fail", "fail"), ("Monitor", "monitor")]:
        w.writerow([label, pfm.get(key, 0)])
    w.writerow([])

    for cat in ["Hydraulics", "Engine", "Electrical"]:
        cad = _category_analytics(cat)
        w.writerow([f"ANALYTICS – {cat} failures over time (Line chart)"])
        w.writerow(["Month", "Count"])
        for r in (cad.get("failures_over_time") or []):
            w.writerow([r.get("month", ""), r.get("count", "")])
        w.writerow([])

    w.writerow(["Chart tips: Select the header row + data rows for a section, then Insert > Chart. Choose Bar, Line, or Pie."])

    return Response(
        content=buf.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename=\"cat-inspect-export-all.csv\"",
            "Content-Type": "text/csv; charset=utf-8",
        },
    )


def _build_export_all_xlsx() -> bytes:
    """Build an Excel workbook with inspection data and embedded charts (bar, line, pie)."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    wb = Workbook()
    inspections = list(MOCK_INSPECTIONS)
    analytics = MOCK_ANALYTICS
    failed_parts = analytics.get("failed_parts") or []
    over_time = analytics.get("inspections_over_time") or []
    pfm = analytics.get("pass_fail_monitor") or {}

    # ---- Sheet 1: Inspections ----
    ws_insp = wb.active
    ws_insp.title = "Inspections"
    ws_insp.append([
        "ID", "Equipment", "Serial", "Customer", "Location", "Date", "Type", "Status", "Inspector",
        "Safety #", "Actions #", "Summary"
    ])
    for i in inspections:
        safety = i.get("safety_findings") or []
        actions = i.get("action_items") or []
        ws_insp.append([
            i.get("id", ""),
            i.get("equipment_model", ""),
            i.get("serial_number", ""),
            i.get("customer", ""),
            i.get("location", ""),
            i.get("date", ""),
            i.get("inspection_type", ""),
            i.get("status", ""),
            i.get("inspector", ""),
            len(safety),
            len(actions),
            (i.get("summary") or "")[:200],
        ])

    n_fp = len(failed_parts)
    n_ot = len(over_time)

    # ---- Sheet 2: Failed Parts + Bar Chart ----
    ws_fp = wb.create_sheet("Failed Parts", 1)
    ws_fp.append(["Category", "Count", "Percentage"])
    for p in failed_parts:
        ws_fp.append([p.get("category", ""), p.get("count", 0), p.get("percentage", 0)])
    if n_fp >= 1:
        try:
            chart_bar = BarChart()
            chart_bar.type = "col"
            chart_bar.title = "Failures by Category"
            chart_bar.y_axis.title = "Count"
            chart_bar.x_axis.title = "Category"
            data = Reference(ws_fp, min_col=2, min_row=1, max_col=2, max_row=n_fp + 1)
            cats = Reference(ws_fp, min_col=1, min_row=2, max_row=n_fp + 1)
            chart_bar.add_data(data, titles_from_data=True)
            chart_bar.set_categories(cats)
            chart_bar.width = 14
            chart_bar.height = 8
            ws_fp.add_chart(chart_bar, "E2")
        except Exception:
            pass

    # ---- Sheet 3: Inspections Over Time + Line Chart ----
    ws_ot = wb.create_sheet("Over Time", 2)
    ws_ot.append(["Month", "Count"])
    for r in over_time:
        ws_ot.append([r.get("month", ""), r.get("count", 0)])
    if n_ot >= 1:
        try:
            chart_line = LineChart()
            chart_line.title = "Inspections Over Time"
            chart_line.y_axis.title = "Count"
            chart_line.x_axis.title = "Month"
            data = Reference(ws_ot, min_col=2, min_row=1, max_col=2, max_row=n_ot + 1)
            cats = Reference(ws_ot, min_col=1, min_row=2, max_row=n_ot + 1)
            chart_line.add_data(data, titles_from_data=True)
            chart_line.set_categories(cats)
            chart_line.width = 14
            chart_line.height = 8
            ws_ot.add_chart(chart_line, "E2")
        except Exception:
            pass

    # ---- Sheet 4: Outcomes + Pie Chart ----
    ws_oc = wb.create_sheet("Outcomes", 3)
    ws_oc.append(["Outcome", "Count"])
    for label, key in [("Pass", "pass"), ("Fail", "fail"), ("Monitor", "monitor")]:
        ws_oc.append([label, pfm.get(key, 0)])
    try:
        chart_pie = PieChart()
        chart_pie.title = "Inspection Outcomes"
        data = Reference(ws_oc, min_col=2, min_row=1, max_col=2, max_row=4)
        cats = Reference(ws_oc, min_col=1, min_row=2, max_row=4)
        chart_pie.add_data(data, titles_from_data=True)
        chart_pie.set_categories(cats)
        chart_pie.width = 10
        chart_pie.height = 8
        ws_oc.add_chart(chart_pie, "E2")
    except Exception:
        pass

    # ---- Sheet 5: Hydraulics Over Time ----
    cad = _category_analytics("Hydraulics")
    fot = cad.get("failures_over_time") or []
    ws_hy = wb.create_sheet("Hydraulics Trend", 4)
    ws_hy.append(["Month", "Failures"])
    for r in fot:
        ws_hy.append([r.get("month", ""), r.get("count", 0)])
    if len(fot) >= 1:
        try:
            chart_hy = LineChart()
            chart_hy.title = "Hydraulics Failures Over Time"
            chart_hy.y_axis.title = "Failures"
            n_hy = len(fot) + 1
            data = Reference(ws_hy, min_col=2, min_row=1, max_col=2, max_row=n_hy)
            cats = Reference(ws_hy, min_col=1, min_row=2, max_row=n_hy)
            chart_hy.add_data(data, titles_from_data=True)
            chart_hy.set_categories(cats)
            chart_hy.width = 12
            chart_hy.height = 7
            ws_hy.add_chart(chart_hy, "E2")
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@api_router.get("/export/all/excel")
async def export_all_excel():
    """Export all inspections and analytics as Excel (.xlsx) with embedded charts (bar, line, pie)."""
    try:
        xlsx_bytes = _build_export_all_xlsx()
    except Exception as e:
        logging.exception("Excel export failed: %s", e)
        raise HTTPException(status_code=500, detail="Excel export failed. Try CSV export instead.")
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=\"cat-inspect-export-all.xlsx\"",
            "Cache-Control": "no-cache, no-store, must-revalidate",
        },
    )


# Document upload for chatbot context (PDF)
@api_router.post("/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    session_id: str = Form(default="inspector-session"),
):
    """Upload a PDF; its text is used as context for the chatbot in this session."""
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed.")
    content = await file.read()
    if len(content) > 15 * 1024 * 1024:  # 15 MB
        raise HTTPException(status_code=400, detail="File too large (max 15 MB).")
    text = extract_text_from_pdf(content)
    if not text:
        raise HTTPException(status_code=400, detail="Could not extract text from PDF. The file may be scanned or corrupted.")
    SESSION_DOCUMENTS[session_id] = {"filename": file.filename or "document.pdf", "text": text}
    preview = text[:300].replace("\n", " ") + ("..." if len(text) > 300 else "")
    return {"success": True, "filename": file.filename, "preview": preview, "char_count": len(text)}

@api_router.delete("/documents")
async def clear_document(session_id: str = "inspector-session"):
    """Clear the uploaded document for this session."""
    if session_id in SESSION_DOCUMENTS:
        del SESSION_DOCUMENTS[session_id]
    return {"success": True}

@api_router.get("/documents/context")
async def get_document_context(session_id: str = "inspector-session"):
    """Return current document info for this session (for UI state)."""
    doc = SESSION_DOCUMENTS.get(session_id)
    if not doc:
        return {"filename": None}
    return {"filename": doc["filename"], "char_count": len(doc["text"])}

# Chat endpoint
@api_router.post("/chat", response_model=ChatResponse)
async def chat(request: ChatRequest):
    """Chat with AI assistant about inspection data and/or uploaded document"""
    try:
        openai_client = get_openai_client()
        session_id = request.session_id or "inspector-session"
        doc = SESSION_DOCUMENTS.get(session_id)

        if doc:
            # Document context: short, plain-text structured responses (no Markdown)
            system_message = """You are Cat Inspect AI Assistant. The user has uploaded a document (PDF) below. Use it to answer their questions.

Response rules:
- Use PLAIN TEXT only. Do not use Markdown: no asterisks for bold (no **), no # headers, no markdown links.
- Structure your reply with clear section labels and line breaks. Format like this exactly:

Key findings:
- Finding one in a short line
- Finding two
- Finding three

Top actions:
- Action one
- Action two

Main risk: One short sentence here.

- Keep each section short (2-5 bullets max). Use a blank line between sections. Bullets can use "- " at the start of the line.
- When the user asks for a chart, graph, breakdown, or visualization: add a single JSON object in your reply with this exact format (numbers can be approximate): {"chart_type": "bar", "title": "Short chart title", "data": [{"category": "Label1", "count": 10}, {"category": "Label2", "count": 5}, ...]}. Use category and count. Up to 6–8 bars is fine. Keep the text reply short.

--- BEGIN UPLOADED DOCUMENT ---
"""
            system_message += doc["text"]
            system_message += "\n--- END UPLOADED DOCUMENT ---"
            max_tokens = 600
        else:
            # Default: inspector context without document
            system_message = """You are Cat Inspect AI Assistant, an expert AI helper for Caterpillar equipment inspectors. 
You have access to the inspector's inspection data and can help with:
- Summarizing inspections
- Identifying recurring failures and patterns
- Providing maintenance recommendations
- Answering questions about equipment

Current inspector: Sriram N.
Recent inspection statistics:
- Total inspections: 224
- Pass rate: 70%
- Most common failures: Hydraulics (35%), Engine (23%), Electrical (17%)
- Equipment types: Excavators, Dozers, Wheel Loaders, Articulated Trucks

Recent inspections summary:
1. CAT 320 Excavator - PASS (Jan 15)
2. CAT D6 Dozer - FAIL (Jan 14) - Hydraulic leak
3. CAT 966 Wheel Loader - MONITOR (Jan 13) - Bucket teeth wear
4. CAT 745 Articulated Truck - PASS (Jan 12)
5. CAT 336 Excavator - In Progress (Jan 11)

When the user asks for a chart, graph, breakdown, or visualization: include a JSON object in your reply with this exact format (numbers can be simple/approximate): {"chart_type": "bar", "title": "Short chart title", "data": [{"category": "Label1", "count": 12}, {"category": "Label2", "count": 8}, ...]}. Use "category" and "count" only. 5–8 bars is enough. Keep your text reply brief.
Use plain text only (no Markdown: no ** for bold, no # headers). Be concise, professional, and helpful. Focus on actionable insights."""
            max_tokens = 500

        response = await openai_client.chat.completions.create(
            model=OPENAI_CHAT_MODEL,
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": request.message}
            ],
            max_tokens=max_tokens
        )
        
        response_text = response.choices[0].message.content
        
        # Extract chart JSON if present (simple bar charts for UI)
        chart_data = _extract_chart_json(response_text)
        return ChatResponse(response=response_text, chart_data=chart_data)
        
    except Exception as e:
        logger.error(f"Chat error: {str(e)}", exc_info=True)
        session_id = request.session_id or "inspector-session"
        doc = SESSION_DOCUMENTS.get(session_id)
        msg_lower = request.message.lower()
        asking_about_doc = doc and ("analyze" in msg_lower or "document" in msg_lower or "summarize" in msg_lower or "insight" in msg_lower or "risk" in msg_lower)

        # If they have a document loaded and asked to analyze it, surface the real error
        if asking_about_doc:
            err_msg = str(e).replace("OpenAI:", "").strip()
            if "invalid_api_key" in err_msg.lower() or "authentication" in err_msg.lower():
                return ChatResponse(response="I couldn't analyze your document: the API key looks invalid. Please check OPENAI_API_KEY in backend/.env and restart the backend.")
            if "context_length" in err_msg.lower() or "maximum context" in err_msg.lower():
                return ChatResponse(response="Your document is too long for one analysis. Try uploading a shorter PDF or ask about a specific section.")
            return ChatResponse(response=f"I couldn't analyze your document. Error: {err_msg[:200]}. Please check your API key and try again.")

        # Fallback response (no document or generic question)
        fallback_responses = {
            "summarize": "Your last inspection was on CAT D6 Dozer (Jan 14) which failed due to a critical hydraulic leak in the main boom cylinder. Immediate repair is recommended before returning the equipment to service.",
            "failures": "Based on your recent inspections, the top recurring failures are: 1) Hydraulics (35%) - mainly hose wear and seal issues, 2) Engine (23%) - oil leaks and filter issues, 3) Electrical (17%) - wiring and sensor problems.",
            "chart": "Here's a breakdown of your failures by category over the last quarter."
        }
        if "summarize" in msg_lower or "last" in msg_lower:
            return ChatResponse(response=fallback_responses["summarize"])
        elif "fail" in msg_lower or "recurring" in msg_lower:
            return ChatResponse(
                response=fallback_responses["failures"],
                chart_data={
                    "chart_type": "bar",
                    "title": "Failures by Category",
                    "data": MOCK_ANALYTICS["failed_parts"]
                }
            )
        elif "chart" in msg_lower or "graph" in msg_lower:
            return ChatResponse(
                response=fallback_responses["chart"],
                chart_data={
                    "chart_type": "bar",
                    "title": "Failures by Category",
                    "data": MOCK_ANALYTICS["failed_parts"]
                }
            )
        else:
            return ChatResponse(response="I can help you with inspection summaries, failure analysis, and equipment recommendations. What would you like to know?")

# AI Vision Analysis - Analyze camera frame for issues
@api_router.post("/ai/vision/analyze")
async def analyze_vision(request: VisionAnalysisRequest):
    """Analyze an image for equipment issues using GPT-4o Vision"""
    try:
        openai_client = get_openai_client()
        
        system_prompt = """You are an expert Caterpillar equipment inspector AI assistant. 
Analyze the image and identify any issues, defects, or safety concerns.

Look for:
- Hydraulic leaks (fluid stains, wet areas, drips)
- Rust and corrosion (orange/brown discoloration, surface pitting)
- Physical damage (dents, cracks, broken parts)
- Wear patterns (worn surfaces, thin materials, degradation)
- Safety hazards (loose parts, missing guards, exposed wiring)
- Part identification (identify visible components)

Respond in this JSON format:
{
    "summary": "Brief 1-2 sentence summary of what you see",
    "findings": [
        {"issue": "description", "severity": "HIGH/MEDIUM/LOW", "location": "where on equipment", "recommendation": "what to do"}
    ],
    "overall_severity": "HIGH/MEDIUM/LOW/NONE",
    "should_alert": true/false (true if HIGH severity found),
    "spoken_response": "A natural spoken sentence to tell the inspector what you found (keep it brief and actionable)"
}

If you don't see any clear issues, still provide a brief assessment.
Be concise but thorough. Focus on actionable findings."""

        response = await openai_client.chat.completions.create(
            model=OPENAI_CHAT_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "Analyze this equipment image for any issues or concerns:"},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{request.image_base64}",
                                "detail": "high"
                            }
                        }
                    ]
                }
            ],
            max_tokens=1000
        )
        
        response_text = response.choices[0].message.content
        
        # Parse JSON response
        import json
        import re
        
        # Try to extract JSON from response
        json_match = re.search(r'\{[\s\S]*\}', response_text)
        if json_match:
            try:
                result = json.loads(json_match.group())
                return {
                    "analysis": result.get("summary", "Analysis complete"),
                    "findings": result.get("findings", []),
                    "severity": result.get("overall_severity", "NONE"),
                    "should_alert": result.get("should_alert", False),
                    "spoken_response": result.get("spoken_response", "I've completed my analysis.")
                }
            except json.JSONDecodeError:
                pass
        
        # Fallback if JSON parsing fails
        return {
            "analysis": response_text[:200] if response_text else "Analysis complete",
            "findings": [],
            "severity": "NONE",
            "should_alert": False,
            "spoken_response": "I've analyzed the image but couldn't identify specific issues."
        }
        
    except Exception as e:
        logger.error(f"Vision analysis error: {str(e)}")
        return {
            "analysis": "Unable to analyze image at this time.",
            "findings": [],
            "severity": "NONE", 
            "should_alert": False,
            "spoken_response": "I'm having trouble analyzing the image right now."
        }

# Text to Speech - Convert AI response to audio
@api_router.post("/ai/tts")
async def text_to_speech(request: TTSRequest):
    """Convert text to speech using OpenAI TTS"""
    try:
        import base64
        
        openai_client = get_openai_client()
        
        response = await openai_client.audio.speech.create(
            model="tts-1",
            voice=request.voice or "alloy",
            input=request.text
        )
        
        # Get audio bytes and convert to base64
        audio_bytes = response.content
        audio_base64 = base64.b64encode(audio_bytes).decode('utf-8')
        
        return {"audio_base64": audio_base64, "format": "mp3"}
        
    except Exception as e:
        logger.error(f"TTS error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"TTS generation failed: {str(e)}")

# Speech to Text - Convert user speech to text
@api_router.post("/ai/stt")
async def speech_to_text(request: STTRequest):
    """Convert speech to text using OpenAI Whisper"""
    try:
        import base64
        import tempfile
        import os as os_module
        
        openai_client = get_openai_client()
        
        # Decode audio and save to temp file
        audio_bytes = base64.b64decode(request.audio_base64)
        
        with tempfile.NamedTemporaryFile(suffix='.webm', delete=False) as temp_file:
            temp_file.write(audio_bytes)
            temp_path = temp_file.name
        
        try:
            # Transcribe using OpenAI Whisper
            with open(temp_path, 'rb') as audio_file:
                response = await openai_client.audio.transcriptions.create(
                    model="whisper-1",
                    file=audio_file
                )
            
            return {"text": response.text, "success": True}
        finally:
            # Clean up temp file
            if os_module.path.exists(temp_path):
                os_module.remove(temp_path)
        
    except Exception as e:
        logger.error(f"STT error: {str(e)}")
        return {"text": "", "success": False, "error": str(e)}
        
    except Exception as e:
        logger.error(f"STT error: {str(e)}")
        return {"text": "", "success": False, "error": str(e)}

# Media Storage Models
class MediaUploadRequest(BaseModel):
    inspection_id: str
    media_type: str  # "photo" or "video"
    data_base64: str
    caption: Optional[str] = None
    timestamp: Optional[str] = None

# Store captured media
INSPECTION_MEDIA = {}

@api_router.post("/inspections/{inspection_id}/media")
async def upload_media(inspection_id: str, request: MediaUploadRequest):
    """Store captured photo or video for an inspection"""
    try:
        import base64
        
        media_id = f"m-{uuid.uuid4().hex[:8]}"
        timestamp = request.timestamp or datetime.now(timezone.utc).strftime("%H:%M:%S")
        
        # Create media record
        media_item = {
            "id": media_id,
            "type": request.media_type,
            "data_base64": request.data_base64,  # Store the actual data
            "timestamp": timestamp,
            "caption": request.caption or f"Captured {request.media_type}",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Store in inspection media dict
        if inspection_id not in INSPECTION_MEDIA:
            INSPECTION_MEDIA[inspection_id] = []
        
        INSPECTION_MEDIA[inspection_id].append(media_item)
        
        logger.info(f"Media saved: {media_id} for inspection {inspection_id}")
        
        return {
            "success": True,
            "media_id": media_id,
            "message": f"{request.media_type.capitalize()} saved successfully"
        }
        
    except Exception as e:
        logger.error(f"Media upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to save media: {str(e)}")

@api_router.get("/inspections/{inspection_id}/media")
async def get_inspection_media(inspection_id: str):
    """Get all media for an inspection"""
    media_list = INSPECTION_MEDIA.get(inspection_id, [])
    
    # Return without the full base64 data for listing
    return [{
        "id": m["id"],
        "type": m["type"],
        "timestamp": m["timestamp"],
        "caption": m["caption"],
        "thumbnail": m["data_base64"][:100] + "..." if len(m.get("data_base64", "")) > 100 else m.get("data_base64", "")
    } for m in media_list]

@api_router.get("/inspections/{inspection_id}/media/{media_id}")
async def get_media_item(inspection_id: str, media_id: str):
    """Get a specific media item with full data"""
    media_list = INSPECTION_MEDIA.get(inspection_id, [])
    
    for m in media_list:
        if m["id"] == media_id:
            return m
    
    raise HTTPException(status_code=404, detail="Media not found")

# Include the router in the main app
app.include_router(api_router)

# Include the realtime router under /api/ai
if realtime_chat:
    app.include_router(realtime_router, prefix="/api/ai")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
